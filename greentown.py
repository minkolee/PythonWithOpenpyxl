# 用于生成给绿城制作模拟报表, 从凭证中分解出一级和二级科目的程序
import tools
import openpyxl


# 填充凭证号填充, 然后检查*号, 只保留指定的*号对应的凭证
def clean_data(file_name: str):
    # 填充凭证号, 获取一个中间表
    file = fill_column(6, file_name)

    ws = openpyxl.load_workbook(file).active

    # 遍历, 将*号对应的凭证号放到一个字典中
    # greentown_entries字典用于记录所有的归属于绿城的凭证号.
    greentown_entries = {}

    # 避免ws.max_row变动, 一开始就取好
    row_number = ws.max_row

    for number in range(1, row_number):
        if ws.cell(row=number, column=7).value and ws.cell(row=number, column=7).value.startswith('*'):
            greentown_entries[ws.cell(row=number, column=6).value] = 1

    print(greentown_entries)

    new_wb = openpyxl.Workbook()

    new_ws = new_wb.active

    current_row = 1

    # 按凭证号复制内容
    for each_number in range(1, row_number):
        # 如果是对应的凭证, 就复制到新表中
        if ws.cell(row=each_number, column=6).value in greentown_entries.keys():
            # 复制凭证号,摘要,科目代码科目名称 借方,贷方到新的工作表中
            new_ws.cell(row=current_row, column=1, value=ws.cell(row=each_number, column=6).value)
            new_ws.cell(row=current_row, column=2, value=ws.cell(row=each_number, column=7).value)
            new_ws.cell(row=current_row, column=3, value=ws.cell(row=each_number, column=8).value)
            new_ws.cell(row=current_row, column=4, value=ws.cell(row=each_number, column=9).value)
            new_ws.cell(row=current_row, column=5, value=ws.cell(row=each_number, column=13).value)
            new_ws.cell(row=current_row, column=6, value=ws.cell(row=each_number, column=14).value)
            current_row += 1
    new_wb.save("提取后的凭证.xlsx")
    return "提取后的凭证.xlsx"


# 获取科目与科目代码字典的函数
def get_subjects(file_name: str):
    result_dict = {}

    ws = tools.open_xlsx_file('科目余额表.xlsx')

    row_max = ws.max_row

    for each in range(1, row_max + 1):
        result_dict[ws.cell(row=each, column=1).value.strip()] = ws.cell(row=each, column=2).value.strip()

    print("组装的字典是: {}".format(result_dict))

    return result_dict


# 向提取后的凭证写入一级科目和二级科目的功能.
def format_entry(file_name: str):
    # 组装字典
    subject_dict = get_subjects('科目余额表.xlsx')

    ws = tools.open_xlsx_file(file_name)
    row_max = ws.max_row

    # 提取一级科目
    for each in range(1, row_max + 1):
        ws.cell(row=each, column=7, value=ws.cell(row=each, column=4).value.split('-')[0].strip())
        if ws.cell(row=each, column=3).value.startswith('1221'):
            ws.cell(row=each, column=8, value=ws.cell(each, column=7).value)
        else:
            ws.cell(row=each, column=8,
                    value=ws.cell(each, column=7).value + '-' + subject_dict.get(ws.cell(row=each, column=3).value))

    ws.parent.save(file_name)

    return file_name


# 统计一级科目和二级科目合计数
def count_result(file_name: str):
    result_dict = {}

    wb = openpyxl.load_workbook(file_name)

    ws = wb.active

    row_max = ws.max_row

    # 处理一级科目合计
    for i in range(1, row_max + 1):
        # 如果已经有这个键, 加上借方数字, 减去贷方数字
        if result_dict.get(ws.cell(row=i, column=7).value):
            result_dict[ws.cell(row=i, column=7).value] = result_dict.get(ws.cell(row=i, column=7).value) + ws.cell(
                row=i, column=5).value - ws.cell(row=i, column=6).value
        else:
            result_dict[ws.cell(row=i, column=7).value] = ws.cell(row=i, column=5).value - ws.cell(row=i,
                                                                                                   column=6).value

    ws1 = wb.create_sheet('一级科目汇总')

    start = 2
    ws1['a1'] = '一级科目'
    ws1['b1'] = '金额'
    for key, value in result_dict.items():
        ws1.cell(row=start, column=1).value = key
        ws1.cell(row=start, column=2).value = value
        start = start + 1

    # 处理二级科目
    result_dict.clear()
    for i in range(1, row_max + 1):
        # 如果已经有这个键, 加上借方数字, 减去贷方数字
        if result_dict.get(ws.cell(row=i, column=8).value):
            result_dict[ws.cell(row=i, column=8).value] = result_dict.get(ws.cell(row=i, column=8).value) + ws.cell(
                row=i, column=5).value - ws.cell(row=i, column=6).value
        else:
            result_dict[ws.cell(row=i, column=8).value] = ws.cell(row=i, column=5).value - ws.cell(row=i,
                                                                                                   column=6).value
    start = 2
    ws2 = wb.create_sheet('二级科目汇总')

    ws2['a1'] = '一级科目'
    ws2['b1'] = '二级科目'
    ws2['c1'] = '金额'

    for key, value in result_dict.items():
        ws2.cell(row=start, column=1).value = key.split('-')[0]
        ws2.cell(row=start, column=2).value = key
        ws2.cell(row=start, column=3).value = value
        start = start + 1

    wb.save(file_name)


# 生成填充后的凭证清单, 保存文件, 然后返回保存的文件名
def fill_column(column: int, worksheet):
    ws = tools.open_xlsx_file(worksheet)

    # 由于文件格式统一, 可以直接设置起点行号是2, 结束的行号是 ws.max-row-1
    start = 2
    end = ws.max_row - 1
    print('start is {}, end is {}'.format(start, end))

    # 然后的思路:
    # 从起点到结尾 2 to end
    # 第一行必定是有值的, 从2开始, 向下寻找是None的单元格, 如果是None就填充. 如果越界或者不是None就停止
    # 停止之后, 要么已经越界, 要么停在下一张凭证的第一行, 更新当前的行数, 然后继续执行同样的循环

    current_row = start

    # 外层的循环, 保证不越界
    while current_row <= end:
        # 说明是凭证的第一行, 用一个变量保存一下这个格子的数据
        current_value = ws.cell(row=current_row, column=column).value
        print('当前凭证处理行号: {}, 凭证号是: {}'.format(current_row, current_value))

        # 从第一行之后的下一行进行填充, 一直填充到下一个不为None或者越界结束.
        # 下一行的位置
        next_row = current_row + 1

        # 从下一行开始, 满足是None而且没有越界, 复制current_value的值到这些格子里
        while (not ws.cell(row=next_row, column=column).value) and next_row <= end:
            print("正在填充的行号是: {} , 值是 {}".format(next_row, current_value))
            ws.cell(row=next_row, column=column, value=current_value)
            next_row = next_row + 1

        # 这个循环结束以后, next_row有两种情况, 第一个停在下一个不是None的位置, 第二, 超过end.
        print('一张凭证处理结束, 新行号: {}'.format(next_row))
        # 所以要进行判断, 大于end了, 直接break,说明处理完毕
        if next_row > end:
            print('行号: {} 超过 {}, 填充结束'.format(next_row, end))
            break
        # 如果没大于end, 此时的next_row指向下一条凭证的第一行
        else:
            # 将当前循环变量设置为next_row, 即从这一行开始继续执行相同的逻辑.
            current_row = next_row
            print('将从 {} 行开始处理下一张凭证'.format(current_row))
            print('-----------------------------------------------------------------------')

    ws.parent.save(worksheet.split('.')[0] + '整理后.xlsx')

    return worksheet.split('.')[0] + '整理后.xlsx'


if __name__ == '__main__':
    # 返回提取后的文件名
    fileName = clean_data('5月凭证.xlsx')

    format_entry(fileName)

    count_result(fileName)
