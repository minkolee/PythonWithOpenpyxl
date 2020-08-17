import openpyxl
import random


def sales_list():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['a1'].value = "销售清单"
    number_of_row = random.randrange(25, 60)
    ws.cell(row=2, column=1, value="产品ID")
    ws.cell(row=2, column=2, value="名称")
    ws.cell(row=2, column=3, value="描述")
    ws.cell(row=2, column=4, value="单价")
    ws.cell(row=2, column=5, value="销售数量")
    ws.cell(row=2, column=6, value="销售金额")

    for i in range(3, number_of_row):
        ws.cell(row=i, column=1, value="IN" + str(i).zfill(4))
        ws.cell(row=i, column=2, value='项目' + str(i))
        ws.cell(row=i, column=3, value='描述' + str(i))
        ws.cell(row=i, column=4, value=random.randrange(14, 60))
        ws.cell(row=i, column=5, value=random.randrange(5, 200))
        ws.cell(row=i, column=6, value=ws.cell(row=i, column=5).value * ws.cell(row=i, column=4).value)

    return ws
