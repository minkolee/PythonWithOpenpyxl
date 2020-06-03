import os
import openpyxl
import random
import decimal
import re


# 返回一个目录内所有的.xlsx文件的路径
def excel_file_list(path: str) -> list:
    result_list = []

    for each_entry in os.scandir(path):
        # 如果文件记录是一个文件夹/目录,
        if each_entry.is_dir():
            result_list += excel_file_list(each_entry.path)

        else:
            if each_entry.path.endswith('.xlsx'):
                result_list.append(each_entry.path)

    return result_list


# 返回一个目录内所有的.xlsx文件的路径
def excel_file_list_iter(path: str) -> list:
    result_list = []

    stack = [path]

    while len(stack) != 0:
        current_dir = stack.pop()
        if os.path.isdir(current_dir):
            for each_entry in os.scandir(current_dir):
                if each_entry.is_dir():
                    stack.append(each_entry.path)
                else:
                    if each_entry.path.endswith('.xlsx'):
                        result_list.append(each_entry.path)

        else:
            if current_dir.path.endswith('.xlsx'):
                result_list.append(current_dir.path)

    return result_list


# 月度化一个工作簿
def monthlize(path: str):
    wb = openpyxl.load_workbook(path)

    for i in range(12):
        ws = wb.copy_worksheet(wb.active)
        ws.title = "{}月{}".format(i + 1, wb.active.title)
        ws.sheet_properties.tabColor = '{0:06X}'.format(random.randint(0, 0xFFFFFF))

    wb.remove(wb.active)
    wb.save(path.split('.')[0] + 'monthly.xlsx')


def transfer_to_decimal(num) -> decimal.Decimal:
    # 如果是一个整数, 就直接进行转换
    if type(num) == int:
        return decimal.Decimal(num)

    # 如果是一个浮点数, 将其转换成2位小数的字符串表示, 然后使用字符串来创建Decimal对象
    elif type(num) == float:
        return decimal.Decimal('{0:.2f}'.format(num))

    # 如果是一个字符串, 需要判断其格式
    elif type(num) == str:

        # 判断字符串是不是一个十进制的小数的表示
        if re.match('[+-]?\\d+(\\.\\d+)?$', num):
            # 字符串是一个十进制的小数表示
            # 判断是否是整数, 如果是整数, 直接通过整数创建Decimal对象
            if num.find('.') == -1:
                return decimal.Decimal(num)

            # 不是整数的情况下
            else:
                split_num = num.split('.')

                if len(split_num[1]) > 2:
                    num_string = split_num[0] + '.' + split_num[1][0:2]
                    if int(split_num[1][2]) >= 5:
                        if num[0] == '-':
                            return decimal.Decimal(num_string) - decimal.Decimal('0.01')
                        else:
                            return decimal.Decimal(num_string) + decimal.Decimal('0.01')
                    else:
                        return decimal.Decimal(num_string)

                elif len(split_num[1]) == 2:
                    num_string = split_num[0] + '.' + split_num[1][0:2]
                    return decimal.Decimal(num_string)

                elif len(split_num[1]) == 1:
                    num_string = split_num[0] + '.' + split_num[1] + '0'
                    return decimal.Decimal(num_string)

                else:
                    raise AttributeError

        else:
            raise AttributeError

    # 不是上述三种类型
    else:
        raise AttributeError


# 打开文件并返回指定名称的工作表, 或者返回活动工作表
def open_xlsx_file(file_name: str, sheet_name=None):
    if sheet_name:
        return openpyxl.load_workbook(file_name)[sheet_name]
    else:
        return openpyxl.load_workbook(file_name).active
