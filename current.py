import random
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, NamedStyle, Side

def sales_list(ws, number_of_row):
    ws['a1'].value = "销售清单"
    ws.cell(row=2, column=1, value="产品ID")
    ws.cell(row=2, column=2, value="名称")
    ws.cell(row=2, column=3, value="描述")
    ws.cell(row=2, column=4, value="单价")
    ws.cell(row=2, column=5, value="销售数量")
    ws.cell(row=2, column=6, value="销售金额")

    for i in range(3, number_of_row + 3):
        ws.cell(row=i, column=1, value="IN" + str(i).zfill(4))
        ws.cell(row=i, column=2, value='项目' + str(i))
        ws.cell(row=i, column=3, value='描述' + str(i))
        ws.cell(row=i, column=4, value=random.randrange(14, 60))
        ws.cell(row=i, column=5, value=random.randrange(5, 200))
        ws.cell(row=i, column=6, value=ws.cell(row=i, column=5).value * ws.cell(row=i, column=4).value)

    return ws

wb = openpyxl.Workbook()
ws = wb.active

sales_list(ws, 20).parent.save("new.xlsx")