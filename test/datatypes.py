import openpyxl


wb = openpyxl.load_workbook('test.xlsx')

ws = wb.active

print(ws['A2'].value, "is {}".format(type(ws['A2'].value)))
print(ws['A3'].value, "is {}".format(type(ws['A3'].value)))
print(ws['B2'].value, "is {}".format(type(ws['B2'].value)))
print(ws['B3'].value, "is {}".format(type(ws['B3'].value)))
print(ws['C2'].value, "is {}".format(type(ws['C2'].value)))
print(ws['C3'].value, "is {}".format(type(ws['C3'].value)))
print(ws['D2'].value, "is {}".format(type(ws['D2'].value)))
print(ws['D3'].value, "is {}".format(type(ws['D3'].value)))

wb.close()


print(ws['D3'].value, "is {}".format(type(ws['D3'].value)))
